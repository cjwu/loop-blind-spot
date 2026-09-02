"""Recover per-problem results from 'All gpt result .xlsx' and join to the
292-problem difficulty/type taxonomy. Produces every table the manuscript
lacks. No new experiments: this reads only artifacts already on disk."""
import openpyxl, os, re, json, statistics as st
from itertools import combinations
from collections import Counter, defaultdict

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(BASE, 'All gpt result .xlsx')
ROOT = os.path.join(BASE, '分類(292題)')
MOD  = [('GPT-4', 5, 6), ('4o', 3, 4), ('4o-mini', 1, 2), ('o1-mini', 7, 8)]
DIF  = ['Beginner(1)', 'Basic(2)', 'Intermediate(3)', 'Hard(4)', 'Expert(5)']


def norm(s):
    s = str(s).strip().lower()
    s = re.sub(r'^exams_', '', s)
    s = re.sub(r'_b$', '', s)
    return re.sub(r'[^a-z0-9]', '', s)


def taxonomy():
    lab = {}
    for t in os.listdir(ROOT):
        tp = os.path.join(ROOT, t)
        if not os.path.isdir(tp):
            continue
        for d in os.listdir(tp):
            dp = os.path.join(tp, d)
            if not os.path.isdir(dp):
                continue
            for prob in os.listdir(dp):
                if os.path.isdir(os.path.join(dp, prob)):
                    lab[norm(prob)] = (t.strip(), d.strip())
    return lab


def load():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    rows = [r for r in wb['工作表1'].iter_rows(min_row=2, values_only=True) if r[0]]
    rec = []
    for r in rows:
        if not any(r[c] for _, c, _ in MOD):
            continue
        e = {'name': str(r[0]).strip(), 'key': norm(r[0])}
        for m, c, n in MOD:
            e[m] = {'pass': bool(r[c]) and 'Pass' in str(r[c]), 'iters': r[n]}
        rec.append(e)
    return wb, rec


def main():
    wb, rec = load()
    lab = taxonomy()
    N = len(rec)
    out = {'n_problems': N, 'n_taxonomy': len(lab)}

    # 1. single-model pass rate + convergence distribution
    single = {}
    for m, _, _ in MOD:
        ok = [e[m]['iters'] for e in rec if e[m]['pass'] and e[m]['iters'] is not None]
        single[m] = {
            'pass': sum(1 for e in rec if e[m]['pass']),
            'rate': round(100 * sum(1 for e in rec if e[m]['pass']) / N, 2),
            'median_iters_of_solved': st.median(ok),
            'mean_iters_of_solved': round(st.mean(ok), 2),
            'solved_at_0': sum(1 for v in ok if v == 0),
            'solved_within_1': sum(1 for v in ok if v <= 1),
        }
    out['single_model'] = single

    # 2. cumulative solve curve (reproduces manuscript Table 1)
    out['solve_curve'] = {
        m: {f'k<={k}': round(100 * sum(1 for e in rec
                                       if e[m]['pass'] and e[m]['iters'] is not None
                                       and e[m]['iters'] <= k) / N, 2)
            for k in (0, 1, 2, 3, 5, 10, 15, 25)}
        for m, _, _ in MOD}

    # 3. ensemble ceiling (union) and uniquely-solved problems
    P = {m: {e['name'] for e in rec if e[m]['pass']} for m, _, _ in MOD}
    union = set.union(*P.values())
    out['ensemble'] = {
        'best_single': max(single, key=lambda m: single[m]['rate']),
        'union_all': round(100 * len(union) / N, 2),
        'solved_by_none': round(100 * (N - len(union)) / N, 2),
        'unique': {m: sorted(P[m] - set.union(*[P[x] for x, _, _ in MOD if x != m]))
                   for m, _, _ in MOD},
    }
    for k in (2, 3):
        best = max(combinations([m for m, _, _ in MOD], k),
                   key=lambda c: len(set.union(*[P[m] for m in c])))
        out['ensemble'][f'best_{k}'] = {
            'models': list(best),
            'rate': round(100 * len(set.union(*[P[m] for m in best])) / N, 2)}

    # 4. realised multi-model runs (cumulative histogram in the 4mix sheet)
    ws = wb['4mix']
    hist = [r for r in ws.iter_rows(min_row=2, max_row=5, values_only=True) if r[6] is not None]
    out['realised_multimodel'] = [
        {'combo': c, 'at_0': h[6], 'within_5': h[7], 'within_15': h[8],
         'final': h[9], 'final_rate': round(100 * h[9] / N, 2)}
        for c, h in zip(['mix 4 model', 'o1+4o mini', 'o1+4'], hist)]

    # 5. difficulty x type breakdown WITH denominators
    cell = defaultdict(lambda: defaultdict(lambda: [0, 0]))
    joined = 0
    for e in rec:
        if e['key'] not in lab:
            continue
        joined += 1
        for m, _, _ in MOD:
            cell[lab[e['key']]][m][1] += 1
            if e[m]['pass']:
                cell[lab[e['key']]][m][0] += 1
    out['joined'] = joined
    out['breakdown'] = {f'{t} | {d}': {m: f'{v[0]}/{v[1]}' for m, v in c.items()}
                        for (t, d), c in sorted(cell.items())}

    # 6. failure concentration by circuit type
    tot, none_c = Counter(), Counter()
    for e in rec:
        if e['key'] not in lab:
            continue
        t = lab[e['key']][0]
        tot[t] += 1
        if not any(e[m]['pass'] for m, _, _ in MOD):
            none_c[t] += 1
    out['unsolved_by_type'] = {t: f'{none_c[t]}/{tot[t]} ({100*none_c[t]/tot[t]:.1f}%)'
                               for t in tot}

    # 7. HEADLINE: repair efficacy by circuit type. Among problems the model
    #    failed on its first attempt, what fraction did the feedback loop recover?
    #    This is a within-model, within-benchmark contrast, so it does not depend
    #    on the difficulty labels, on model diversity, or on the benchmark being
    #    standard.
    repair = {}
    for m, ci, ni in MOD:
        by_type = defaultdict(lambda: [0, 0])   # [recovered, first-attempt failures]
        for e in rec:
            if e['key'] not in lab:
                continue
            it = e[m]['iters']
            if it is None:
                continue
            if e[m]['pass'] and it == 0:        # solved zero-shot, not a failure
                continue
            t = lab[e['key']][0]
            by_type[t][1] += 1
            if e[m]['pass']:
                by_type[t][0] += 1
        repair[m] = {t: {'recovered': v[0], 'first_attempt_failures': v[1],
                         'efficacy': round(100 * v[0] / v[1], 1) if v[1] else None}
                     for t, v in by_type.items()}
    pooled = defaultdict(lambda: [0, 0])
    for m in repair:
        for t, v in repair[m].items():
            pooled[t][0] += v['recovered']
            pooled[t][1] += v['first_attempt_failures']
    out['repair_efficacy'] = {'per_model': repair, 'pooled': {
        t: {'recovered': v[0], 'first_attempt_failures': v[1],
            'efficacy': round(100 * v[0] / v[1], 1),
            'ci95_pp': round(196 * (v[0] / v[1] * (1 - v[0] / v[1]) / v[1]) ** 0.5, 1)}
        for t, v in pooled.items()}}

    # 8. three-way outcome class per problem-model pair, by circuit type
    cls = defaultdict(Counter)
    for e in rec:
        if e['key'] not in lab:
            continue
        t = lab[e['key']][0]
        for m, _, _ in MOD:
            it = e[m]['iters']
            if it is None:
                continue
            cls[t]['zeroshot' if (e[m]['pass'] and it == 0)
                   else ('repaired' if e[m]['pass'] else 'never')] += 1
    out['outcome_classes'] = {t: dict(c) for t, c in cls.items()}

    dst = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'recovered.json')
    with open(dst, 'w') as f:
        json.dump(out, f, indent=2, ensure_ascii=False)
    print('wrote', dst)
    print(json.dumps({k: out[k] for k in
                      ('n_problems', 'joined', 'repair_efficacy', 'outcome_classes')},
                     indent=2, ensure_ascii=False))


if __name__ == '__main__':
    main()
